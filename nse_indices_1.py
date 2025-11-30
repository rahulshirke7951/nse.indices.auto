"""
nse_indices_1.py

1) Reads config from Json/nse_broad.json
2) Fetches NSE index data via yfinance
3) Writes:
   - nse_indices_1_raw.xlsx        (raw workbook)
   - nse_indices_1_dashboard.xlsx  (formatted workbook)
   - data/nse_indices_1_latest.csv (summary for Sheets/Apps Script)

Designed for GitHub Actions automation (no Colab, no Drive).
"""

import os
import json
from datetime import date

import pandas as pd
import yfinance as yf

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# === BASE PATHS ===

BASE = os.path.dirname(os.path.abspath(__file__))

JSON_CFG = os.path.join(BASE, "Json", "nse_broad.json")

RAW_XL = os.path.join(BASE, "nse_indices_1_raw.xlsx")
OUT_XL = os.path.join(BASE, "nse_indices_1_dashboard.xlsx")

CACHE_DIR = os.path.join(BASE, "cache")
DATA_DIR = os.path.join(BASE, "data")

os.makedirs(os.path.join(BASE, "Json"), exist_ok=True)
os.makedirs(CACHE_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

# === LOAD CONFIG ===

if not os.path.exists(JSON_CFG):
    raise FileNotFoundError(f"Config file missing: {JSON_CFG}")

with open(JSON_CFG, "r") as f:
    cfg = json.load(f)

INDICES = cfg.get("indices", {})
START_DATE = cfg.get("start_date")
END_DATE = cfg.get("end_date")

today = date.today()
if START_DATE and END_DATE:
    start_date = pd.to_datetime(START_DATE).date()
    end_date = pd.to_datetime(END_DATE).date()
else:
    start_date = today.replace(day=1)
    end_date = today

print("Date range:", start_date, "→", end_date)
print("Indices:", list(INDICES.keys()))

# === FETCH DATA ===

data_dict = {}
for name, symbol in INDICES.items():
    try:
        df = yf.Ticker(symbol).history(
            start=start_date,
            end=(pd.to_datetime(end_date) + pd.Timedelta(days=1)).date()
        )
        if df is not None and not df.empty:
            data_dict[name] = df["Close"].copy()

            csvp = os.path.join(CACHE_DIR, f"{symbol.replace('^','caret_')}.csv")
            try:
                df.to_csv(csvp)
            except Exception:
                pass

            print(f"Fetched: {name} ({symbol}) rows={len(df)}")
        else:
            print(f"No data: {name} ({symbol})")
    except Exception as e:
        print(f"Error fetching {name}: {e}")

if not data_dict:
    raise RuntimeError("No data fetched for any index.")

# === PROCESS DATA ===

df_close = pd.DataFrame(data_dict)
df_close.index = pd.to_datetime(df_close.index).tz_localize(None)
df_close = (
    df_close
    .sort_index(ascending=True)
    .dropna(axis=1, how="all")
    .ffill()
    .bfill()
)

df_pct_mtd = ((df_close / df_close.iloc[0]) - 1) * 100
df_pct_mtd = df_pct_mtd.round(2)

df_pct_dod = df_close.pct_change() * 100
df_pct_dod = df_pct_dod.round(2)

mtd_series = df_pct_mtd.iloc[-1].copy()
summary = (
    pd.DataFrame({"MTD % Change": mtd_series})
    .sort_values(by="MTD % Change", ascending=False)
)

daily_summary = []
for dt in df_pct_dod.sort_index(ascending=False).index:
    row = df_pct_dod.loc[dt].dropna()
    if len(row) > 0:
        top_g = row.nlargest(3)
        top_l = row.nsmallest(3)
        daily_summary.append({
            "Date": dt.strftime("%d-%b-%y"),
            "Top 3 Gainers": ", ".join(top_g.index),
            "Top 3 Losers": ", ".join(top_l.index)
        })
daily_summary_df = pd.DataFrame(daily_summary)

streaks = {}
for col in df_close.columns:
    diffs = df_close[col].diff().fillna(0)

    max_win = max_loss = cur = 0
    for v in (diffs > 0):
        cur = cur + 1 if v else 0
        max_win = max(max_win, cur)

    cur = 0
    for v in (diffs < 0):
        cur = cur + 1 if v else 0
        max_loss = max(max_loss, cur)

    streaks[col] = {
        "Longest Win Streak": int(max_win),
        "Longest Lose Streak": int(max_loss)
    }

streaks_df = (
    pd.DataFrame.from_dict(streaks, orient="index")
    .reset_index()
    .rename(columns={"index": "Index"})
)

avg_change = summary["MTD % Change"].mean()
gain_cnt = int((summary["MTD % Change"] > 0).sum())
loss_cnt = int((summary["MTD % Change"] < 0).sum())
best = summary.index[0] if len(summary) else ""
worst = summary.index[-1] if len(summary) else ""
mood = "Bullish" if avg_change > 0 else ("Bearish" if avg_change < 0 else "Neutral")

market_overview = pd.DataFrame({
    "Key Insight": [
        "Date Range", "Average Market Change (%)", "Total Gainers",
        "Total Losers", "Top Performer", "Bottom Performer", "Market Mood"
    ],
    "Value": [
        f"{start_date.strftime('%d-%b-%Y')} → {end_date.strftime('%d-%b-%Y')}",
        f"{round(avg_change, 2)}%",
        f"{gain_cnt}",
        f"{loss_cnt}",
        best,
        worst,
        mood
    ]
})

# === WRITE RAW WORKBOOK (same sheet names as before) ===

with pd.ExcelWriter(RAW_XL, engine="openpyxl") as w:
    df_close.sort_index(ascending=False).to_excel(w, sheet_name="Index Close")
    df_pct_mtd.sort_index(ascending=False).to_excel(w, sheet_name="MTD %")
    df_pct_dod.sort_index(ascending=False).to_excel(w, sheet_name="Day over Day %")
    summary.to_excel(w, sheet_name="Summary")
    daily_summary_df.to_excel(w, sheet_name="Daily Movers", index=False)
    streaks_df.to_excel(w, sheet_name="Streaks", index=False)
    market_overview.to_excel(w, sheet_name="Market Overview", index=False)

print("✅ Raw workbook saved:", RAW_XL)

# === WRITE CSV FOR APPS SCRIPT ===

csv_path = os.path.join(DATA_DIR, "nse_indices_1_latest.csv")
summary.to_csv(csv_path)
print("✅ Latest summary CSV saved:", csv_path)

# === FORMATTING WORKBOOK (dashboard) ===

if not os.path.exists(RAW_XL):
    raise FileNotFoundError(f"Raw workbook missing: {RAW_XL}")

wb = load_workbook(RAW_XL)

# ---- COMMON STYLES ----
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_NO_WRAP = Alignment(horizontal="center", vertical="center")
HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
NORMAL_FONT = Font(bold=False)
DATE_FMT = "dd-mmm-yy"

# ---- HELPERS ----
def remove_all_borders(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.border = None

def set_header_style(ws):
    for c in ws[1]:
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.fill = HEADER_FILL
        c.border = None

def set_col_widths(ws, widths):
    if isinstance(widths, tuple):
        ws.column_dimensions[get_column_letter(1)].width = widths[0]
        for i in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = widths[1]

def format_date_column(ws, col_idx=1, fmt=DATE_FMT, align=CENTER_NO_WRAP):
    for r in range(2, ws.max_row + 1):
        c = ws.cell(r, col_idx)
        c.number_format = fmt
        c.font = NORMAL_FONT
        c.alignment = align

def add_mtd_dod_heatmap(ws, min_row=2, min_col=2):
    """
    Red (bad) → Yellow (neutral) → Green (good) 3‑color scale
    """
    if ws.max_column < min_col or ws.max_row < min_row:
        return

    start_col_letter = get_column_letter(min_col)
    end_col_letter = get_column_letter(ws.max_column)
    rng = f"{start_col_letter}{min_row}:{end_col_letter}{ws.max_row}"

    rule = ColorScaleRule(
        start_type="num", start_value=-2, start_color="FFBE5014",     # red
        mid_type="num",  mid_value=0,  mid_color="FFFFF59D",         # yellow
        end_type="num",  end_value=2,  end_color="FF92D050"          # green
    )
    ws.conditional_formatting.add(rng, rule)

# ---- APPLY GLOBAL HEADER + NO BORDERS ----
for ws in wb.worksheets:
    remove_all_borders(ws)
    set_header_style(ws)

# ---- COLUMN WIDTHS (same as original intentions) ----
set_col_widths(wb["Index Close"], (16, 14))
set_col_widths(wb["MTD %"], (16, 12))
set_col_widths(wb["Day over Day %"], (16, 12))
set_col_widths(wb["Summary"], (22, 16))
set_col_widths(wb["Daily Movers"], (16, 30))
set_col_widths(wb["Streaks"], (20, 14))
set_col_widths(wb["Market Overview"], (28, 28))

# ---- INDEX CLOSE: date + integer close ----
ws = wb["Index Close"]
format_date_column(ws, col_idx=1)

for col in range(2, ws.max_column + 1):
    for r in range(2, ws.max_row + 1):
        ws.cell(r, col).number_format = "0"

# ---- MTD %: date + heatmap ----
ws = wb["MTD %"]
format_date_column(ws, col_idx=1)
add_mtd_dod_heatmap(ws, min_row=2, min_col=2)

# ---- DoD%: rename + date + heatmap ----
ws = wb["Day over Day %"]
format_date_column(ws, col_idx=1)
add_mtd_dod_heatmap(ws, min_row=2, min_col=2)
ws.title = "DoD%"  # keep dashboard sheet naming you used earlier

# ---- SUMMARY: header text & normal index font ----
ws = wb["Summary"]
ws["A1"].value = "Index"
for cell in ws["A"][1:]:
    cell.font = NORMAL_FONT

# ---- DAILY MOVERS: proper date + wide text columns ----
ws = wb["Daily Movers"]
for r in range(2, ws.max_row + 1):
    cell = ws.cell(r, 1)
    val = cell.value
    if isinstance(val, str):
        dt = pd.to_datetime(val, errors="coerce", dayfirst=True)
        if not pd.isna(dt):
            cell.value = dt
    cell.number_format = DATE_FMT
    cell.font = NORMAL_FONT
    cell.alignment = CENTER_NO_WRAP

ws.column_dimensions["B"].width = 60
ws.column_dimensions["C"].width = 60

# ---- STREAKS: numeric columns as integer ----
ws = wb["Streaks"]
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 20

for header_cell in ws[1]:
    if header_cell.value in ("Longest Win Streak", "Longest Lose Streak"):
        col_letter = header_cell.column_letter
        for cell in ws[col_letter][1:]:
            cell.number_format = "0"

# ---- MARKET OVERVIEW → Overview + bold left column ----
ws = wb["Market Overview"]
ws.title = "Overview"
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 36

for cell in ws["A"]:
    cell.font = HEADER_FONT

# ---- SHEET ORDER (preserve names & new DoD%) ----
desired = ["Overview", "Summary", "Index Close", "MTD %", "DoD%", "Daily Movers", "Streaks"]
wb._sheets = [wb[s] for s in desired if s in wb.sheetnames]

wb.save(OUT_XL)
print("✅ Dashboard workbook saved:", OUT_XL)
print("🎉 nse_indices_1 pipeline completed.")
